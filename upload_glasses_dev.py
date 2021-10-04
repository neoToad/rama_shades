import os
import django
from openpyxl import Workbook, load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "rama_shades.settings")
django.setup()

directory = "available_inventory.xlsx"
parent_dir = "C:/Users/colin/PycharmProjects/who_cutie_scraper"
workbook_path = os.path.join(parent_dir, directory)

from store.models import Item, ItemImage

wb = load_workbook(workbook_path)

ws1 = wb['product_details']
row_count = ws1.max_row

objs = []

# Bulk Create
for row_idx in range(2, 3):
    slug = '-'.join(ws1.cell(row=row_idx, column=2).value.split(' '))[:30]
    print(f'C:/Users/colin/PycharmProjects/who_cutie_scraper/items/item-{ws1.cell(row=row_idx, column=1).value}/main_img.png')

    if not Item.objects.filter(slug=slug).exists():
        objs.append(
            Item(
                title=ws1.cell(row=row_idx, column=2).value[:99],
                description=ws1.cell(row=row_idx, column=3).value,
                additional_info=ws1.cell(row=row_idx, column=4).value,
                image=f'images/{ws1.cell(row=row_idx, column=1).value}_main_img.png',
                image_hover=f'images/{ws1.cell(row=row_idx, column=1).value}_hover_img.png',
                price=ws1.cell(row=row_idx, column=10).value,
                discount_price=ws1.cell(row=row_idx, column=11).value,
                slug=slug,
                item_id=ws1.cell(row=row_idx, column=1).value,
                 )
        )

    else:
        item = Item.objects.get(slug=slug)
        img_01, created = ItemImage.objects.get_or_create(
            image=f'images/{ws1.cell(row=row_idx, column=1).value}_img_1.png',
            defaults={
                'item': item
            }
        )
        img_02, created = ItemImage.objects.get_or_create(
            image=f'images/{ws1.cell(row=row_idx, column=1).value}_img_2.png',
            defaults={
                'item': item
            }
        )
        img_03, created = ItemImage.objects.get_or_create(
            image=f'images/{ws1.cell(row=row_idx, column=1).value}_img_3.png',
            defaults={
                'item': item
            }
        )


Item.objects.bulk_create(objs)

